#!/usr/bin/env python3
"""Build the audit log and the human-review report queue.

Inputs:
  data/reviews.jsonl          (review metadata: url, asin, rating, date)
  data/classifications.jsonl  (per-review decisions)

Outputs:
  output/audit_log.jsonl   EVERY review checked, with its full decision + a
                           run timestamp. Append-only audit trail.
  output/report_queue.xlsx  ONLY reviews with is_violation == true AND
                           confidence >= confidence_threshold. One row per
                           flagged review, with an empty "approve (y/n)" column
                           for a human to fill in.
"""

from __future__ import annotations

import sys
from datetime import datetime, timezone

from _common import append_jsonl, load_config, read_jsonl, resolve

# Where each violation type should be reported. Amazon has no report API, so
# these are the human-facing channels.
REPORT_CHANNELS: dict[str, str] = {
    "promotional": "Brand Registry > Report a Violation",
    "fake_incentivized": "Brand Registry > Report a Violation",
    "plagiarized": "Brand Registry > Report a Violation",
    "off_topic": 'Review "Report abuse" link',
    "profanity": 'Review "Report abuse" link',
    "hate_harassment": 'Review "Report abuse" link',
    "private_info": 'Review "Report abuse" link',
    "illegal_dangerous": 'Review "Report abuse" link',
}

QUEUE_COLUMNS = [
    "review_id",
    "review_url",
    "asin",
    "rating",
    "date",
    "violation_type",
    "confidence",
    "evidence_quote",
    "matched_guideline",
    "report_justification",
    "recommended_channel",
    "approve (y/n)",
]


def draft_justification(violation_type: str, guideline: str, quote: str) -> str:
    """A neutral, factual report justification a human can paste as-is."""
    readable = violation_type.replace("_", " ")
    text = (
        f"This review appears to violate Amazon's community guideline on "
        f"{readable}"
    )
    if guideline:
        text += f" ({guideline})"
    text += "."
    if quote:
        text += f' Specifically, the review states: "{quote}".'
    text += " Requesting review for removal under the applicable guideline."
    return text


def main() -> None:
    config = load_config()
    threshold = float(config.get("confidence_threshold", 0.8))

    reviews = {r["review_id"]: r for r in read_jsonl(resolve(config["paths"]["reviews"]))}
    classifications = read_jsonl(resolve(config["paths"]["classifications"]))
    if not classifications:
        sys.exit("error: no classifications found — run classify.py first")

    run_ts = datetime.now(timezone.utc).isoformat()
    audit_path = resolve(config["paths"]["audit_log"])

    flagged: list[dict] = []
    for decision in classifications:
        rid = decision["review_id"]
        review = reviews.get(rid, {})

        # Audit log: every review checked, full decision, timestamp.
        append_jsonl(
            audit_path,
            {
                "logged_at": run_ts,
                "review_id": rid,
                "asin": decision.get("asin", review.get("asin", "")),
                "rating": review.get("rating"),
                "url": review.get("url", ""),
                "is_violation": decision.get("is_violation", False),
                "violation_type": decision.get("violation_type", "none"),
                "confidence": decision.get("confidence", 0.0),
                "matched_guideline": decision.get("matched_guideline", ""),
                "evidence_quote": decision.get("evidence_quote", ""),
                "justification": decision.get("justification", ""),
                "classified_at": decision.get("classified_at", ""),
                "model": decision.get("model", ""),
                "error": decision.get("error"),
                "threshold": threshold,
            },
        )

        # Report queue: only confident, real violations.
        confidence = float(decision.get("confidence", 0.0) or 0.0)
        if decision.get("is_violation") and not decision.get("error") and confidence >= threshold:
            vtype = decision.get("violation_type", "none")
            flagged.append(
                {
                    "review_id": rid,
                    "review_url": review.get("url", ""),
                    "asin": decision.get("asin", review.get("asin", "")),
                    "rating": review.get("rating", ""),
                    "date": review.get("date", ""),
                    "violation_type": vtype,
                    "confidence": round(confidence, 3),
                    "evidence_quote": decision.get("evidence_quote", ""),
                    "matched_guideline": decision.get("matched_guideline", ""),
                    "report_justification": draft_justification(
                        vtype,
                        decision.get("matched_guideline", ""),
                        decision.get("evidence_quote", ""),
                    ),
                    "recommended_channel": REPORT_CHANNELS.get(
                        vtype, "Brand Registry > Report a Violation"
                    ),
                    "approve (y/n)": "",
                }
            )

    _write_xlsx(resolve(config["paths"]["report_queue"]), flagged)
    print(
        f"audit log appended ({len(classifications)} decisions) -> {audit_path}\n"
        f"report queue: {len(flagged)} flagged reviews "
        f"(confidence >= {threshold}) -> {resolve(config['paths']['report_queue'])}"
    )


def _write_xlsx(path, rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("error: the 'openpyxl' package is required (pip install -r requirements.txt)")

    wb = Workbook()
    ws = wb.active
    ws.title = "report_queue"

    ws.append(QUEUE_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(col, "") for col in QUEUE_COLUMNS])

    # Reasonable, readable column widths.
    widths = {
        "review_id": 18, "review_url": 40, "asin": 12, "rating": 8, "date": 14,
        "violation_type": 18, "confidence": 11, "evidence_quote": 50,
        "matched_guideline": 28, "report_justification": 60,
        "recommended_channel": 34, "approve (y/n)": 14,
    }
    for idx, col in enumerate(QUEUE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)

    wb.save(path)


if __name__ == "__main__":
    main()
