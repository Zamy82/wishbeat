#!/usr/bin/env python3
"""Process human-approved rows from the report queue, one at a time.

Reads output/report_queue.xlsx and acts ONLY on rows whose "approve (y/n)"
column is "y" (case-insensitive). Submission is always human-confirmed and
never bulk — Amazon has no official "report review" API.

Modes:
  manual  (default) Print the review URL, the report channel + steps, and the
                    drafted justification to paste. The human submits in the
                    browser, then confirms here so it is logged.
  assisted          Intended to run inside Claude Code with the "Claude in
                    Chrome" integration: Claude opens the report flow and
                    pre-fills the justification, then STOPS for explicit human
                    confirmation before each submit. This script drives the
                    per-item confirmation gate and the audit log; it never
                    auto-submits on its own.

Idempotent: review_ids already recorded as submitted in
output/report_actions.jsonl are skipped.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone

from _common import append_jsonl, load_config, read_jsonl, resolve


def _read_queue(path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("error: the 'openpyxl' package is required (pip install -r requirements.txt)")
    if not path.exists():
        sys.exit(f"error: report queue not found at {path} — run build_report.py first")

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]


def _print_packet(row: dict, mode: str) -> None:
    print("\n" + "=" * 72)
    print(f"Review ID : {row.get('review_id', '')}")
    print(f"ASIN      : {row.get('asin', '')}")
    print(f"Violation : {row.get('violation_type', '')}  (confidence {row.get('confidence', '')})")
    print(f"URL       : {row.get('review_url', '')}")
    print(f"Channel   : {row.get('recommended_channel', '')}")
    print(f"Guideline : {row.get('matched_guideline', '')}")
    print(f"Evidence  : {row.get('evidence_quote', '')}")
    print("\nReport steps:")
    print("  1. Open the review URL above.")
    print(f"  2. Use: {row.get('recommended_channel', '')}.")
    print("  3. Paste the justification below into the report form.")
    if mode == "assisted":
        print("  (assisted) Claude in Chrome can open this page and pre-fill the")
        print("             justification. Review it, then confirm to submit.")
    print("\nJustification to paste:")
    print(f"  {row.get('report_justification', '')}")
    print("=" * 72)


def main() -> None:
    parser = argparse.ArgumentParser(description="Submit approved review reports, one at a time.")
    parser.add_argument("--mode", choices=["manual", "assisted"], default="manual")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List the approved, not-yet-submitted rows without prompting.",
    )
    args = parser.parse_args()

    config = load_config()
    queue = _read_queue(resolve(config["paths"]["report_queue"]))
    actions_path = resolve(config["paths"]["report_actions"])

    already_done = {
        a["review_id"]
        for a in read_jsonl(actions_path)
        if a.get("action") == "submitted"
    }

    approved = [
        row
        for row in queue
        if str(row.get("approve (y/n)", "")).strip().lower() in ("y", "yes")
        and str(row.get("review_id", "")).strip() not in already_done
    ]

    if not approved:
        print("No approved rows pending. Mark rows with 'y' in the "
              "approve (y/n) column of report_queue.xlsx first.")
        return

    print(f"{len(approved)} approved row(s) pending in {args.mode} mode.")

    if args.dry_run:
        for row in approved:
            print(f"  - {row.get('review_id', '')}  {row.get('violation_type', '')}  {row.get('review_url', '')}")
        return

    for row in approved:
        _print_packet(row, args.mode)
        prompt = (
            "Confirm this report was submitted? [yes = log as submitted / "
            "skip / quit]: "
        )
        try:
            answer = input(prompt).strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\nStopped. No further rows processed.")
            break

        ts = datetime.now(timezone.utc).isoformat()
        if answer in ("yes", "y"):
            append_jsonl(
                actions_path,
                {
                    "review_id": row.get("review_id", ""),
                    "asin": row.get("asin", ""),
                    "violation_type": row.get("violation_type", ""),
                    "channel": row.get("recommended_channel", ""),
                    "mode": args.mode,
                    "action": "submitted",
                    "confirmed_at": ts,
                },
            )
            print("  logged as submitted.")
        elif answer in ("quit", "q"):
            print("Stopped. No further rows processed.")
            break
        else:
            append_jsonl(
                actions_path,
                {
                    "review_id": row.get("review_id", ""),
                    "mode": args.mode,
                    "action": "skipped",
                    "confirmed_at": ts,
                },
            )
            print("  skipped (not logged as submitted).")

    print(f"\nDone. Actions recorded in {actions_path}")


if __name__ == "__main__":
    main()
