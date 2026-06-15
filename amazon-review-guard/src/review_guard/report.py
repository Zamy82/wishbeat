"""Build the audit log (every decision) and the report queue (only confident,
real violations) as an .xlsx with an empty approve column for a human."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .config import Config
from .io_utils import append_jsonl, read_jsonl

# Where each violation type should be reported. Amazon has no report API.
REPORT_CHANNELS: dict[str, str] = {
    "promotional": "Brand Registry > Report a Violation",
    "fake_incentivized": "Brand Registry > Report a Violation",
    "plagiarized": "Brand Registry > Report a Violation",
    "off_topic": 'Review "Report abuse" link',
    "profanity": 'Review "Report abuse" link',
    "hate_harassment": 'Review "Report abuse" link',
    "private_info": 'Review "Report abuse" link',
    "illegal_dangerous": 'Review "Report abuse" link',
}
DEFAULT_CHANNEL = "Brand Registry > Report a Violation"

QUEUE_COLUMNS = [
    "review_id",
    "review_url",
    "asin",
    "rating",
    "date",
    "violation_type",
    "confidence",
    "evidence_quote",
    "matched_guideline",
    "report_justification",
    "recommended_channel",
    "approve (y/n)",
]


def draft_justification(violation_type: str, guideline: str, quote: str) -> str:
    """A neutral, factual report justification a human can paste as-is."""
    readable = violation_type.replace("_", " ")
    text = f"This review appears to violate Amazon's community guideline on {readable}"
    if guideline:
        text += f" ({guideline})"
    text += "."
    if quote:
        text += f' Specifically, the review states: "{quote}".'
    text += " Requesting review for removal under the applicable guideline."
    return text


def build_rows(config: Config) -> tuple[list[dict[str, Any]], int]:
    """Return (queue_rows, num_decisions) and append the audit log.

    queue_rows contains only is_violation==true, no error, confidence>=threshold.
    """
    reviews = {r["review_id"]: r for r in read_jsonl(config.path("reviews"))}
    decisions = read_jsonl(config.path("classifications"))
    if not decisions:
        raise SystemExit("no classifications found — run 'review-guard classify' first")

    threshold = config.confidence_threshold
    run_ts = datetime.now(timezone.utc).isoformat()
    audit_path = config.path("audit_log")

    rows: list[dict[str, Any]] = []
    for decision in decisions:
        rid = decision["review_id"]
        review = reviews.get(rid, {})

        append_jsonl(
            audit_path,
            {
                "logged_at": run_ts,
                "review_id": rid,
                "asin": decision.get("asin", review.get("asin", "")),
                "rating": review.get("rating"),
                "url": review.get("url", ""),
                "is_violation": decision.get("is_violation", False),
                "violation_type": decision.get("violation_type", "none"),
                "confidence": decision.get("confidence", 0.0),
                "matched_guideline": decision.get("matched_guideline", ""),
                "evidence_quote": decision.get("evidence_quote", ""),
                "justification": decision.get("justification", ""),
                "classified_at": decision.get("classified_at", ""),
                "model": decision.get("model", ""),
                "error": decision.get("error"),
                "threshold": threshold,
            },
        )

        confidence = float(decision.get("confidence", 0.0) or 0.0)
        if decision.get("is_violation") and not decision.get("error") and confidence >= threshold:
            vtype = decision.get("violation_type", "none")
            rows.append(
                {
                    "review_id": rid,
                    "review_url": review.get("url", ""),
                    "asin": decision.get("asin", review.get("asin", "")),
                    "rating": review.get("rating", ""),
                    "date": review.get("date", ""),
                    "violation_type": vtype,
                    "confidence": round(confidence, 3),
                    "evidence_quote": decision.get("evidence_quote", ""),
                    "matched_guideline": decision.get("matched_guideline", ""),
                    "report_justification": draft_justification(
                        vtype,
                        decision.get("matched_guideline", ""),
                        decision.get("evidence_quote", ""),
                    ),
                    "recommended_channel": REPORT_CHANNELS.get(vtype, DEFAULT_CHANNEL),
                    "approve (y/n)": "",
                }
            )
    return rows, len(decisions)


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - import guard
        raise SystemExit("the 'openpyxl' package is required (pip install -e .)") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "report_queue"
    ws.append(QUEUE_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(col, "") for col in QUEUE_COLUMNS])

    widths = {
        "review_id": 18, "review_url": 40, "asin": 12, "rating": 8, "date": 14,
        "violation_type": 18, "confidence": 11, "evidence_quote": 50,
        "matched_guideline": 28, "report_justification": 60,
        "recommended_channel": 34, "approve (y/n)": 14,
    }
    for idx, col in enumerate(QUEUE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)
    wb.save(path)


def run_report(config: Config) -> dict[str, Any]:
    rows, num_decisions = build_rows(config)
    queue_path = config.path("report_queue")
    write_xlsx(queue_path, rows)
    return {
        "decisions": num_decisions,
        "flagged": len(rows),
        "queue_path": queue_path,
        "audit_path": config.path("audit_log"),
    }
