"""Process human-approved rows from the report queue, one at a time.

Acts ONLY on rows whose "approve (y/n)" column is y/yes. Submission is always
human-confirmed and never bulk — Amazon has no official report-review API.
Idempotent: review_ids already logged as submitted are skipped.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from .config import Config
from .io_utils import append_jsonl, read_jsonl


def read_queue(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - import guard
        raise SystemExit("the 'openpyxl' package is required (pip install -e .)") from exc
    if not path.exists():
        raise SystemExit(f"report queue not found at {path} — run 'review-guard report' first")

    ws = load_workbook(path, read_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]


def _approved(row: dict[str, Any]) -> bool:
    return str(row.get("approve (y/n)", "")).strip().lower() in ("y", "yes")


def pending_rows(queue: list[dict[str, Any]], done_ids: set[str]) -> list[dict[str, Any]]:
    """Approved rows not yet logged as submitted."""
    return [
        row for row in queue
        if _approved(row) and str(row.get("review_id", "")).strip() not in done_ids
    ]


def format_packet(row: dict[str, Any], mode: str) -> str:
    lines = [
        "=" * 72,
        f"Review ID : {row.get('review_id', '')}",
        f"ASIN      : {row.get('asin', '')}",
        f"Violation : {row.get('violation_type', '')}  (confidence {row.get('confidence', '')})",
        f"URL       : {row.get('review_url', '')}",
        f"Channel   : {row.get('recommended_channel', '')}",
        f"Guideline : {row.get('matched_guideline', '')}",
        f"Evidence  : {row.get('evidence_quote', '')}",
        "",
        "Report steps:",
        "  1. Open the review URL above.",
        f"  2. Use: {row.get('recommended_channel', '')}.",
        "  3. Paste the justification below into the report form.",
    ]
    if mode == "assisted":
        lines += [
            "  (assisted) Claude in Chrome can open this page and pre-fill the",
            "             justification. Review it, then confirm to submit.",
        ]
    lines += [
        "",
        "Justification to paste:",
        f"  {row.get('report_justification', '')}",
        "=" * 72,
    ]
    return "\n".join(lines)


def run_submit(
    config: Config,
    *,
    mode: str = "manual",
    dry_run: bool = False,
    input_fn: Callable[[str], str] = input,
    print_fn: Callable[..., None] = print,
) -> dict[str, int]:
    """Drive the per-item confirmation gate. Returns counts of submitted/skipped."""
    queue = read_queue(config.path("report_queue"))
    actions_path = config.path("report_actions")
    done_ids = {
        a["review_id"] for a in read_jsonl(actions_path) if a.get("action") == "submitted"
    }
    pending = pending_rows(queue, done_ids)

    if not pending:
        print_fn(
            "No approved rows pending. Mark rows with 'y' in the "
            "approve (y/n) column of the report queue first."
        )
        return {"submitted": 0, "skipped": 0}

    print_fn(f"{len(pending)} approved row(s) pending in {mode} mode.")

    if dry_run:
        for row in pending:
            print_fn(f"  - {row.get('review_id', '')}  {row.get('violation_type', '')}  {row.get('review_url', '')}")
        return {"submitted": 0, "skipped": 0}

    submitted = skipped = 0
    for row in pending:
        print_fn("\n" + format_packet(row, mode))
        try:
            answer = input_fn(
                "Confirm this report was submitted? [yes = log as submitted / skip / quit]: "
            ).strip().lower()
        except (EOFError, KeyboardInterrupt):
            print_fn("\nStopped. No further rows processed.")
            break

        ts = datetime.now(timezone.utc).isoformat()
        if answer in ("yes", "y"):
            append_jsonl(
                actions_path,
                {
                    "review_id": row.get("review_id", ""),
                    "asin": row.get("asin", ""),
                    "violation_type": row.get("violation_type", ""),
                    "channel": row.get("recommended_channel", ""),
                    "mode": mode,
                    "action": "submitted",
                    "confirmed_at": ts,
                },
            )
            submitted += 1
            print_fn("  logged as submitted.")
        elif answer in ("quit", "q"):
            print_fn("Stopped. No further rows processed.")
            break
        else:
            append_jsonl(
                actions_path,
                {"review_id": row.get("review_id", ""), "mode": mode, "action": "skipped", "confirmed_at": ts},
            )
            skipped += 1
            print_fn("  skipped (not logged as submitted).")

    return {"submitted": submitted, "skipped": skipped}
