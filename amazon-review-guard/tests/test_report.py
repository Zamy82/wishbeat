from openpyxl import load_workbook

from review_guard.io_utils import read_jsonl, write_jsonl
from review_guard.report import draft_justification, run_report


def test_draft_justification_is_neutral_and_quotes_evidence():
    text = draft_justification("promotional", "No URLs allowed", "www.x.com")
    assert "promotional" in text
    assert "No URLs allowed" in text
    assert '"www.x.com"' in text


def _seed(project):
    write_jsonl(
        project.path("reviews"),
        [
            {"review_id": "R1", "asin": "B000000000", "rating": 5, "date": "2026-05-01", "url": "https://x/R1"},
            {"review_id": "R2", "asin": "B000000000", "rating": 1, "date": "2026-05-02", "url": "https://x/R2"},
            {"review_id": "R3", "asin": "B000000001", "rating": 2, "date": "2026-05-03", "url": "https://x/R3"},
        ],
    )
    write_jsonl(
        project.path("classifications"),
        [
            # honest negative-but-not-a-violation -> excluded
            {"review_id": "R1", "asin": "B000000000", "is_violation": False, "violation_type": "none",
             "confidence": 0.99, "evidence_quote": "", "matched_guideline": "", "error": None},
            # real violation, high confidence -> included
            {"review_id": "R2", "asin": "B000000000", "is_violation": True, "violation_type": "promotional",
             "confidence": 0.95, "evidence_quote": "www.x.com", "matched_guideline": "No URLs", "error": None},
            # violation but below threshold -> excluded
            {"review_id": "R3", "asin": "B000000001", "is_violation": True, "violation_type": "off_topic",
             "confidence": 0.40, "evidence_quote": "late delivery", "matched_guideline": "Product only", "error": None},
        ],
    )


def test_run_report_filters_by_threshold_and_writes_queue(project):
    _seed(project)
    result = run_report(project)
    assert result["decisions"] == 3
    assert result["flagged"] == 1  # only R2 (>= 0.80 and is_violation)

    wb = load_workbook(result["queue_path"])
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert "approve (y/n)" in header
    assert len(rows) == 1
    row = dict(zip(header, rows[0]))
    assert row["review_id"] == "R2"
    assert row["violation_type"] == "promotional"
    assert row["recommended_channel"] == "Brand Registry > Report a Violation"
    assert row["approve (y/n)"] in (None, "")


def test_audit_log_records_every_decision(project):
    _seed(project)
    run_report(project)
    audit = read_jsonl(project.path("audit_log"))
    assert len(audit) == 3  # every reviewed decision logged, not just violations
    assert all("logged_at" in a and "threshold" in a for a in audit)


def test_errored_decision_is_never_flagged(project):
    write_jsonl(project.path("reviews"), [{"review_id": "R1", "asin": "B000000000", "url": "u"}])
    write_jsonl(
        project.path("classifications"),
        [{"review_id": "R1", "asin": "B000000000", "is_violation": True, "violation_type": "promotional",
          "confidence": 0.99, "evidence_quote": "x", "matched_guideline": "g", "error": "Boom: down"}],
    )
    result = run_report(project)
    assert result["flagged"] == 0
